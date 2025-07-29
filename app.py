import streamlit as st
import pandas as pd
import os
from pathlib import Path
import tempfile

# Nombres de columnas finales (las que se guardan - lado derecho del mapeo)
NOMBRES_COLUMNAS = [
    "RFC",
    "CLIENTE",
    "DESPACHO",
    "CODIGO",
    "REFERENCIA",
    "CONCEPTO",
    "CANTIDAD",
    "IMPORTE",
    "IMPUESTO",
]

# Mapeo de etiquetas a buscar en el Excel → nombre de columna final
MAPEO_COLUMNAS = {
    "RFC": "RFC",
    "CLIENTE": "CLIENTE",
    "CUENTA CONTABLE": "CODIGO",
    "REFERENCIA": "REFERENCIA",
    "Descripción": "CONCEPTO",
    "CANTIDAD": "CANTIDAD",
    "Precio": "IMPORTE",
}


def encontrar_fila_rfc(df, fila_inicio=0):
    """Encuentra la fila donde aparece RFC en la columna A, comenzando desde fila_inicio"""
    for i in range(fila_inicio, len(df)):
        if pd.notna(df.iloc[i, 0]) and str(df.iloc[i, 0]).strip().upper() == "RFC":
            return i
    return None


def extraer_info_cliente(df):
    """Extrae información del cliente de la sección de encabezado"""
    info_cliente = {}

    # Verificar si CLIENTE está en A3 (índice de fila 2)
    if len(df) > 2 and pd.notna(df.iloc[2, 0]):
        if str(df.iloc[2, 0]).strip().upper() == "CLIENTE":
            # Buscar características del cliente hacia abajo
            for i in range(
                3, min(len(df), 20)
            ):  # Verificar máximo las próximas 17 filas
                if pd.notna(df.iloc[i, 0]):
                    clave = str(df.iloc[i, 0]).strip()
                    valor = (
                        str(df.iloc[i, 1]).strip() if pd.notna(df.iloc[i, 1]) else ""
                    )

                    if clave.upper() in ["RFC", "CLIENTE", "CUENTA CONTABLE"]:
                        info_cliente[clave] = valor

                    if clave.upper() == "RFC":
                        break

    return info_cliente


def encontrar_valor_por_etiqueta(datos_fila, etiqueta):
    """Busca una etiqueta específica en la fila y devuelve el valor que está en la siguiente columna"""
    for i, celda in enumerate(datos_fila):
        if pd.notna(celda) and str(celda).strip().upper() == etiqueta.upper():
            # Si encontramos la etiqueta, devolver el valor de la siguiente columna
            if i + 1 < len(datos_fila) and pd.notna(datos_fila[i + 1]):
                return str(datos_fila[i + 1]).strip()
    return ""


def encontrar_columnas_por_nombre(datos_fila):
    """Encuentra las posiciones de las columnas basándose en los nombres de las etiquetas"""
    posiciones_columnas = {}

    for i, celda in enumerate(datos_fila):
        if pd.notna(celda):
            nombre_celda = str(celda).strip()
            # Mapear los nombres encontrados a nuestros nombres de columna
            if nombre_celda.upper() == "RFC":
                posiciones_columnas["RFC"] = i
            elif nombre_celda.upper() == "CLIENTE":
                posiciones_columnas["CLIENTE"] = i
            elif nombre_celda.upper() == "CUENTA CONTABLE":
                posiciones_columnas["CODIGO"] = i
            elif nombre_celda.upper() == "REFERENCIA":
                posiciones_columnas["REFERENCIA"] = i
            elif (
                nombre_celda == "Descripción" or "Descripción" in nombre_celda
            ):  # Buscar con tilde y más flexible
                posiciones_columnas["CONCEPTO"] = i
            elif (
                "CANTIDAD" in nombre_celda.upper()
            ):  # Más flexible: captura "Cantidad" y "Cantidad STU"
                posiciones_columnas["CANTIDAD"] = i
                print(
                    f"🔍 DEBUG: Encontrada columna CANTIDAD como '{nombre_celda}' en posición {i}"
                )
            elif (
                nombre_celda == "Precio" or "Precio" in nombre_celda
            ):  # Más flexible también
                posiciones_columnas["IMPORTE"] = i

    return posiciones_columnas


def extraer_datos_de_fila(datos_fila, posiciones_columnas):
    """Extrae los datos de una fila usando las posiciones de columnas identificadas"""
    datos_concepto = {}

    for nombre_columna, posicion in posiciones_columnas.items():
        if posicion < len(datos_fila) and pd.notna(datos_fila[posicion]):
            valor = str(datos_fila[posicion]).strip()
            if valor:  # Solo agregar si no está vacío
                datos_concepto[nombre_columna] = valor

    return datos_concepto


def es_fila_totales_factura(datos_fila):
    """Detecta si una fila contiene totales de factura (SUBTOTAL, IVA, TOTAL)"""
    palabras_totales = [
        "SUBTOTAL",
        "IVA",
        "TOTAL",
        "SUMA",
        "IMPORTE TOTAL",
        "TOTAL FACTURA",
    ]

    # Convertir la fila a texto para análisis
    texto_fila = []
    for celda in datos_fila:
        if pd.notna(celda):
            texto_fila.append(str(celda).strip().upper())

    # Buscar palabras clave de totales
    tiene_palabras_totales = any(
        any(palabra in texto for palabra in palabras_totales) for texto in texto_fila
    )

    if tiene_palabras_totales:
        # Contar números en la fila (los totales suelen tener varios números)
        numeros_encontrados = 0
        for texto in texto_fila:
            try:
                # Limpiar el texto y ver si es un número
                numero_limpio = (
                    texto.replace(",", "")
                    .replace("$", "")
                    .replace("%", "")
                    .replace("(", "")
                    .replace(")", "")
                )
                float(numero_limpio)
                numeros_encontrados += 1
            except:
                pass

        # Si tiene palabras de totales Y números, probablemente es fila de totales
        return numeros_encontrados >= 1

    return False


def es_fila_titulos_columna(datos_concepto):
    """Detecta si un concepto extraído contiene títulos de columna en lugar de datos reales"""
    if not datos_concepto:
        return False

    # Palabras que indican que es una fila de títulos, no datos
    palabras_titulos = [
        "RFC",
        "CLIENTE",
        "CODIGO",
        "REFERENCIA",
        "CONCEPTO",
        "CANTIDAD",
        "IMPORTE",
        "DESCRIPCION",
        "DESCRIPCIÓN",
        "PRECIO",
        "CANTIDAD STU",
        "NO. FACTURA",
    ]

    # Contar cuántas columnas contienen palabras de títulos
    coincidencias_titulos = 0
    total_columnas_con_datos = 0

    for valor in datos_concepto.values():
        if valor and str(valor).strip():
            total_columnas_con_datos += 1
            valor_upper = str(valor).strip().upper()

            # Verificar coincidencias exactas o parciales con títulos
            for palabra_titulo in palabras_titulos:
                if (
                    valor_upper == palabra_titulo
                    or palabra_titulo in valor_upper
                    or valor_upper in palabra_titulo
                ):
                    coincidencias_titulos += 1
                    break

    # Si más del 50% de las columnas con datos son títulos, es una fila de títulos
    if total_columnas_con_datos > 0:
        porcentaje_titulos = coincidencias_titulos / total_columnas_con_datos
        es_titulo = porcentaje_titulos > 0.5

        if es_titulo:
            print(
                f"🏷️ DEBUG: Fila de títulos detectada - {coincidencias_titulos}/{total_columnas_con_datos} coincidencias"
            )
            print(f"   Datos: {dict(datos_concepto)}")

        return es_titulo

    return False


def extraer_facturas_de_hoja(df, nombre_hoja):
    """Extrae todas las facturas de una sola hoja de Excel"""
    facturas = []
    fila_actual = 0

    # Extraer información del cliente primero
    info_cliente = extraer_info_cliente(df)

    while fila_actual < len(df):
        # Encontrar el próximo RFC en la columna A
        fila_rfc = encontrar_fila_rfc(df, fila_actual)

        if fila_rfc is None:
            break

        # En la fila donde encontramos RFC, identificar las posiciones de las columnas
        datos_fila_rfc = df.iloc[fila_rfc]
        posiciones_columnas = encontrar_columnas_por_nombre(datos_fila_rfc)

        # Leer conceptos en las filas siguientes usando las posiciones identificadas
        conceptos = []
        for i in range(fila_rfc + 1, len(df)):
            # Verificar si la fila está vacía (todos NaN o cadenas vacías)
            datos_fila = df.iloc[i]

            # MEJORADO: Verificar si es fila vacía O fila de totales
            es_fila_vacia = datos_fila.isna().all() or all(
                str(celda).strip() == "" for celda in datos_fila.fillna("")
            )
            es_fila_total = es_fila_totales_factura(datos_fila)

            if es_fila_vacia:
                print(f"📄 DEBUG: Fila vacía encontrada en {i + 1}, terminando factura")
                fila_actual = i + 1
                break
            elif es_fila_total:
                print(
                    f"📊 DEBUG: Fila de totales encontrada en {i + 1}: {[str(celda)[:20] for celda in datos_fila[:6] if pd.notna(celda)]}"
                )
                print(f"   Terminando factura y saltando filas de totales...")

                # Saltar todas las filas de totales consecutivas
                j = i + 1
                while j < len(df):
                    siguiente_fila = df.iloc[j]
                    if (
                        siguiente_fila.isna().all()
                        or all(
                            str(celda).strip() == ""
                            for celda in siguiente_fila.fillna("")
                        )
                        or not es_fila_totales_factura(siguiente_fila)
                    ):
                        break
                    print(f"   Saltando fila de totales adicional en {j + 1}")
                    j += 1

                fila_actual = j
                break
            else:
                # Esta fila contiene datos de concepto
                datos_concepto = extraer_datos_de_fila(datos_fila, posiciones_columnas)

                # NUEVO: Verificar si es fila de títulos antes de procesarla
                if es_fila_titulos_columna(datos_concepto):
                    print(f"🏷️ Saltando fila de títulos en fila {i + 1}")
                    continue

                # Reasignar el CODIGO basado en si contiene "Servicio" (case insensitive)
                if "CONCEPTO" in datos_concepto:
                    codigo_original = str(datos_concepto["CONCEPTO"]).lower()
                    if "servicio" in codigo_original:
                        datos_concepto["CODIGO"] = "76101500"  # Código para servicios
                    else:
                        datos_concepto["CODIGO"] = (
                            "42281522"  # Código para no servicios
                        )

                # Solo agregar si encontramos al menos algunos datos y no son títulos
                if datos_concepto:
                    conceptos.append(datos_concepto)
        else:
            # Si llegamos al final de la hoja sin encontrar una fila vacía
            fila_actual = len(df)

        # Crear objeto factura
        if conceptos:
            factura = {
                "nombre_hoja": nombre_hoja,
                "info_cliente": info_cliente,
                "fila_rfc": fila_rfc + 1,  # +1 para indexación basada en 1
                "conceptos": conceptos,
                "total_conceptos": len(conceptos),
            }
            facturas.append(factura)
            print(
                f"✅ Factura creada con {len(conceptos)} conceptos (RFC en fila {fila_rfc + 1})"
            )

    return facturas, info_cliente


def extraer_todas_facturas(archivo_excel):
    """Extrae facturas de todas las hojas del archivo Excel"""
    todas_facturas = []
    resumenes_hojas = {}

    for nombre_hoja in archivo_excel.sheet_names:
        try:
            df = pd.read_excel(archivo_excel, sheet_name=nombre_hoja, header=None)
            facturas, info_cliente = extraer_facturas_de_hoja(df, nombre_hoja)

            # Almacenar resumen de esta hoja
            resumenes_hojas[nombre_hoja] = {
                "cantidad_facturas": len(facturas),
                "info_cliente": info_cliente,
                "filas_hoja": df.shape[0],
                "columnas_hoja": df.shape[1],
            }

            # Agregar todas las facturas de esta hoja
            todas_facturas.extend(facturas)

        except Exception as e:
            st.error(f"Error procesando la hoja '{nombre_hoja}': {str(e)}")
            resumenes_hojas[nombre_hoja] = {
                "cantidad_facturas": 0,
                "info_cliente": {},
                "error": str(e),
            }

    return todas_facturas, resumenes_hojas


def mostrar_resumen_hojas(resumenes_hojas, todas_facturas=None):
    """Muestra un resumen de todas las hojas con información actualizada de las facturas"""
    st.subheader("📋 Resumen de Hojas")

    datos_resumen = []

    # Si tenemos facturas reales, usar esa información (más precisa)
    facturas_por_hoja = {}
    if todas_facturas:
        for factura in todas_facturas:
            nombre_hoja = factura["nombre_hoja"]
            if nombre_hoja not in facturas_por_hoja:
                facturas_por_hoja[nombre_hoja] = []
            facturas_por_hoja[nombre_hoja].append(factura)

    for nombre_hoja, resumen in resumenes_hojas.items():
        if "error" in resumen:
            datos_resumen.append(
                {
                    "Hoja": nombre_hoja,
                    "Facturas": "❌ Error",
                    "Cliente": "Error",
                    "RFC": "Error",
                    "Tamaño": "Error",
                }
            )
        else:
            # Usar datos reales de las facturas si están disponibles
            if todas_facturas and nombre_hoja in facturas_por_hoja:
                facturas_hoja = facturas_por_hoja[nombre_hoja]

                # Obtener cliente y RFC del primer concepto de la primera factura
                cliente_real = "No encontrado"
                rfc_real = "No encontrado"

                if facturas_hoja and facturas_hoja[0]["conceptos"]:
                    primer_concepto = facturas_hoja[0]["conceptos"][0]
                    cliente_real = primer_concepto.get("CLIENTE", "No encontrado")
                    rfc_real = primer_concepto.get("RFC", "No encontrado")

                datos_resumen.append(
                    {
                        "Hoja": nombre_hoja,
                        "Facturas": len(facturas_hoja),
                        "Cliente": cliente_real,
                        "RFC": rfc_real,
                        "Tamaño": f"{resumen['filas_hoja']}×{resumen['columnas_hoja']}",
                    }
                )
            else:
                # Fallback a la información del resumen original
                nombre_cliente = resumen["info_cliente"].get("CLIENTE", "No encontrado")
                rfc_cliente = resumen["info_cliente"].get("RFC", "No encontrado")
                datos_resumen.append(
                    {
                        "Hoja": nombre_hoja,
                        "Facturas": resumen["cantidad_facturas"],
                        "Cliente": nombre_cliente,
                        "RFC": rfc_cliente,
                        "Tamaño": f"{resumen['filas_hoja']}×{resumen['columnas_hoja']}",
                    }
                )

    df_resumen = pd.DataFrame(datos_resumen)
    st.dataframe(df_resumen, use_container_width=True)


def consolidar_facturas_para_excel(todas_facturas):
    """Consolida todas las facturas en un DataFrame listo para exportar con numeración"""
    filas_consolidadas = []
    numero_factura = 1

    for factura in todas_facturas:
        # Todos los conceptos de esta factura tendrán el mismo número
        for concepto in factura["conceptos"]:
            fila = {
                "No. Factura": numero_factura,
                "Hoja Origen": factura["nombre_hoja"],
                "DESPACHO": "MIDESPACHO",  # Valor fijo para todos los conceptos
                "RFC": concepto.get("RFC", ""),
                "CLIENTE": concepto.get("CLIENTE", ""),
                "CODIGO": concepto.get("CODIGO", ""),
                "REFERENCIA": concepto.get("REFERENCIA", ""),
                "CONCEPTO": concepto.get("CONCEPTO", ""),
                "CANTIDAD": concepto.get("CANTIDAD", ""),
                "IMPORTE": concepto.get("IMPORTE", ""),
                "IMPUESTO": "IVA16",  # Valor fijo para todos los conceptos
                # Columnas adicionales del Template SAT (vacías por ahora)
                "FECHA": "",
                "MONEDA": "MXN",
                "TIPO_CAMBIO": "1.00",
                "SUBTOTAL": "",
                "IVA": "",
                "TOTAL": "",
                "FORMA_PAGO": "",
                "METODO_PAGO": "",
                "USO_CFDI": "",
            }
            filas_consolidadas.append(fila)

        # Incrementar número de factura para la siguiente
        numero_factura += 1

    return pd.DataFrame(filas_consolidadas)


def cargar_template_sat():
    """Carga el Template SAT.xlsx existente"""
    template_path = Path("hanovaexcel/Template SAT.xlsx")

    if not template_path.exists():
        st.error(
            "No se encontró el archivo 'Template SAT.xlsx' en la carpeta hanovaexcel"
        )
        return None, None

    try:
        # Cargar el archivo Excel manteniendo el formato
        from openpyxl import load_workbook

        wb = load_workbook(template_path)
        ws = wb.active

        # También cargar con pandas para análisis
        df = pd.read_excel(template_path, header=None)

        return wb, df
    except Exception as e:
        st.error(
            "❌ No se pudo cargar el archivo Template SAT. Verifica que esté en la carpeta correcta."
        )
        print(f"❌ DEBUG: Error cargando Template SAT: {str(e)}")
        return None, None


def encontrar_fila_titulos_template(df):
    """Encuentra la fila de títulos en el Template SAT de manera robusta"""
    # Buscar por la celda C16 como referencia inicial
    try:
        # Verificar si hay algo en la fila 15 (índice 15, que es la fila 16 en Excel)
        if len(df) > 15:
            fila_16 = df.iloc[15]  # Fila 16 en Excel (índice 15)

            # Verificar si la columna C (índice 2) tiene datos
            if pd.notna(fila_16.iloc[2]) and str(fila_16.iloc[2]).strip():
                st.info(
                    f"Encontrados títulos en fila 16, columna C: '{fila_16.iloc[2]}'"
                )
                return 15  # Índice 15 = fila 16 en Excel
    except:
        pass

    # Método alternativo: buscar por nombres comunes de columnas
    for i in range(min(30, len(df))):  # Buscar en las primeras 30 filas
        fila = df.iloc[i]
        fila_str = " ".join([str(cell).upper() for cell in fila if pd.notna(cell)])

        # Buscar palabras clave que indiquen que es una fila de títulos
        palabras_clave = [
            "FACTURA",
            "RFC",
            "CLIENTE",
            "CONCEPTO",
            "IMPORTE",
            "CANTIDAD",
        ]
        coincidencias = sum(1 for palabra in palabras_clave if palabra in fila_str)

        if coincidencias >= 3:  # Si encuentra al menos 3 palabras clave
            st.info(
                f"Encontrados títulos en fila {i + 1} por coincidencias de palabras clave"
            )
            return i

    # Si no encuentra nada, usar fila 15 como default
    st.warning(
        "No se pudieron encontrar títulos automáticamente, usando fila 16 por defecto"
    )
    return 15


def obtener_mapeo_columnas_template(df, fila_titulos):
    """Obtiene el mapeo de columnas del Template SAT"""
    if fila_titulos >= len(df):
        return {}

    titulos = df.iloc[fila_titulos]
    mapeo = {}

    for i, titulo in enumerate(titulos):
        if pd.notna(titulo):
            titulo_str = str(titulo).strip().upper()

            # Mapear títulos del template a nuestros datos
            if "FACTURA" in titulo_str or "NO." in titulo_str:
                mapeo["No. Factura"] = i
            elif "DESPACHO" in titulo_str:
                mapeo["DESPACHO"] = i
            elif "RFC" in titulo_str:
                mapeo["RFC"] = i
            elif "CLIENTE" in titulo_str:
                mapeo["CLIENTE"] = i
            elif "CODIGO" in titulo_str or "CÓDIGO" in titulo_str:
                mapeo["CODIGO"] = i
            elif "REFERENCIA" in titulo_str:
                mapeo["REFERENCIA"] = i
            elif (
                "CONCEPTO" in titulo_str
                or "DESCRIPCION" in titulo_str
                or "DESCRIPCIÓN" in titulo_str
            ):
                mapeo["CONCEPTO"] = i
            elif "CANTIDAD" in titulo_str:
                mapeo["CANTIDAD"] = i
            elif "IMPORTE" in titulo_str or "PRECIO" in titulo_str:
                mapeo["IMPORTE"] = i
            elif "IMPUESTO" in titulo_str:
                mapeo["IMPUESTO"] = i
            elif "FECHA" in titulo_str:
                mapeo["FECHA"] = i
            elif "MONEDA" in titulo_str:
                mapeo["MONEDA"] = i
            elif "SUBTOTAL" in titulo_str:
                mapeo["SUBTOTAL"] = i
            elif "IVA" in titulo_str:
                mapeo["IVA"] = i
            elif "TOTAL" in titulo_str:
                mapeo["TOTAL"] = i

    return mapeo


def copiar_formato_fila(ws, fila_origen, fila_destino, max_columnas=20):
    """Copia el formato de una fila origen a una fila destino de manera más robusta"""
    try:
        from copy import copy

        for col in range(1, max_columnas + 1):
            celda_origen = ws.cell(row=fila_origen, column=col)
            celda_destino = ws.cell(row=fila_destino, column=col)

            # Copiar todos los atributos de formato
            try:
                celda_destino.font = copy(celda_origen.font)
            except:
                pass

            try:
                celda_destino.border = copy(celda_origen.border)
            except:
                pass

            try:
                celda_destino.fill = copy(celda_origen.fill)
            except:
                pass

            try:
                celda_destino.number_format = celda_origen.number_format
            except:
                pass

            try:
                celda_destino.protection = copy(celda_origen.protection)
            except:
                pass

            try:
                celda_destino.alignment = copy(celda_origen.alignment)
            except:
                pass

            # Copiar alto de fila si es diferente del default
            try:
                if ws.row_dimensions[fila_origen].height:
                    ws.row_dimensions[fila_destino].height = ws.row_dimensions[
                        fila_origen
                    ].height
            except:
                pass

    except Exception as e:
        st.warning(
            f"No se pudo copiar formato de fila {fila_origen} a fila {fila_destino}: {str(e)}"
        )


def llenar_template_sat_con_datos(
    wb, df_template, todas_facturas, fila_titulos, mapeo_columnas
):
    """Llena el Template SAT con los datos de las facturas, copiando formato de filas existentes"""
    ws = wb.active

    # Generar datos consolidados
    df_consolidado = consolidar_facturas_para_excel(todas_facturas)

    # Comenzar a llenar desde la fila siguiente a los títulos
    fila_inicio_datos = (
        fila_titulos + 2
    )  # +1 para la siguiente fila, +1 porque Excel usa índice base 1
    total_filas_datos = len(df_consolidado)

    # Usar la primera fila después de títulos como origen de formato (fila_titulos + 1 en Excel)
    fila_formato_origen = (
        fila_titulos + 1 + 1
    )  # +1 para siguiente fila después de títulos, +1 para Excel

    # Verificar si la fila de origen tiene formato visible pero limpiar cualquier dato existente
    try:
        celda_test = ws.cell(row=fila_formato_origen, column=1)
        tiene_formato = (
            celda_test.font.name != "Calibri"
            or celda_test.font.size != 11
            or str(celda_test.border.left.style) != "None"
            or str(celda_test.fill.fill_type) != "None"
        )
        # DEBUG: Mover a consola
        print(
            f"🔍 DEBUG: Usando fila {fila_formato_origen} como origen de formato ({'con' if tiene_formato else 'sin'} formato visible)"
        )

        # IMPORTANTE: Limpiar cualquier contenido de la fila de formato origen que pueda ser títulos
        for col in range(1, 21):  # Limpiar hasta 20 columnas
            celda = ws.cell(row=fila_formato_origen, column=col)
            if celda.value:  # Si tiene algún valor (posiblemente títulos)
                celda.value = None  # Limpiar el valor pero mantener formato

    except Exception as e:
        print(f"🔍 DEBUG: Error detectando formato origen: {str(e)}")

    # Copiar formato a todas las filas donde insertaremos datos
    print(
        f"📋 DEBUG: Copiando formato de fila {fila_formato_origen} a {total_filas_datos} filas de datos..."
    )
    filas_copiadas = 0

    for i in range(total_filas_datos):
        fila_destino = fila_inicio_datos + i
        copiar_formato_fila(ws, fila_formato_origen, fila_destino)
        filas_copiadas += 1

    print(f"✅ DEBUG: Formato copiado a {filas_copiadas} filas")

    # Ahora insertar los datos (asegurándonos de que no insertamos títulos)
    datos_insertados = 0
    filas_saltadas = 0

    try:
        # Iterar sobre el DataFrame de manera más segura
        for index in df_consolidado.index:
            fila_datos = df_consolidado.loc[index]
            fila_excel = fila_inicio_datos + datos_insertados

            # Verificar que no estamos insertando una fila de títulos
            valores_fila = []
            for col in df_consolidado.columns:
                valor = fila_datos[col]
                if pd.notna(valor):
                    valores_fila.append(str(valor))

            es_fila_titulos = any(
                valor.upper()
                in [
                    "RFC",
                    "CLIENTE",
                    "CODIGO",
                    "REFERENCIA",
                    "CONCEPTO",
                    "CANTIDAD",
                    "IMPORTE",
                    "NO. FACTURA",
                ]
                for valor in valores_fila
            )

            if es_fila_titulos:
                filas_saltadas += 1
                print(f"🚫 DEBUG FILA SALTADA #{filas_saltadas}:")
                print(f"   Índice: {index}")
                print(f"   Valores: {valores_fila}")
                print(f"   Fila completa: {dict(fila_datos)}")
                print(f"   Razón: Contiene palabras de títulos")
                print("   ---")
                continue

            # Debugging de datos válidos
            if datos_insertados < 3:  # Solo para las primeras 3 filas para no saturar
                print(f"✅ DEBUG FILA VÁLIDA #{datos_insertados + 1}:")
                print(f"   Índice: {index}")
                print(f"   Fila Excel destino: {fila_excel}")
                print(f"   Datos a insertar: {dict(fila_datos)}")
                print("   ---")

            # Llenar cada columna según el mapeo
            for nombre_columna, col_index in mapeo_columnas.items():
                if nombre_columna in df_consolidado.columns:
                    try:
                        valor = fila_datos[nombre_columna]
                        if pd.notna(valor) and str(valor).strip():
                            # Escribir en la celda (Excel usa índice base 1)
                            ws.cell(
                                row=fila_excel, column=col_index + 1, value=str(valor)
                            )
                    except Exception as e:
                        print(
                            f"❌ DEBUG ERROR: Error insertando {nombre_columna}: {str(e)}"
                        )
                        continue

            datos_insertados += 1

    except Exception as e:
        print(f"❌ DEBUG ERROR GENERAL: {str(e)}")
        st.error(
            "❌ Hubo un problema al procesar los datos. Revisa la consola para más detalles."
        )
        return None

    print(f"📊 DEBUG RESUMEN:")
    print(f"   Total filas procesadas: {len(df_consolidado)}")
    print(f"   Filas saltadas: {filas_saltadas}")
    print(f"   Filas insertadas: {datos_insertados}")
    print("   " + "=" * 50)

    if datos_insertados > 0:
        st.success(
            f"✅ Se procesaron exitosamente {datos_insertados} conceptos de facturación"
        )
        if filas_saltadas > 0:
            st.info("ℹ️ Se omitieron algunos registros duplicados o incorrectos")
    else:
        st.warning(
            "⚠️ No se pudieron procesar los datos. Verifica el formato de tus archivos."
        )

    return wb


def mostrar_excel_consolidado(todas_facturas, resumenes_hojas):
    """Muestra el Excel consolidado final con todas las facturas"""
    if not todas_facturas:
        st.warning("📋 No hay facturas que mostrar. Sube archivos Excel primero.")
        return

    # Importar BytesIO para los buffers de descarga
    from io import BytesIO

    st.subheader("📊 Excel Consolidado - Vista Previa")

    # Generar DataFrame consolidado
    df_consolidado = consolidar_facturas_para_excel(todas_facturas)

    # Cargar template SAT para verificar si los datos están listos
    wb, df_template = cargar_template_sat()
    template_listo = False
    mapeo_columnas = {}

    if wb is not None and df_template is not None:
        # Encontrar fila de títulos y mapeo (sin mostrar análisis)
        fila_titulos = encontrar_fila_titulos_template(df_template)
        mapeo_columnas = obtener_mapeo_columnas_template(df_template, fila_titulos)
        template_listo = bool(mapeo_columnas)

        # DEBUG: Mover a consola
        print(f"🔍 DEBUG TEMPLATE SAT:")
        print(f"   Fila de títulos detectada: {fila_titulos + 1} (Excel)")
        if mapeo_columnas:
            print(f"   Columnas detectadas en Template SAT:")
            for nombre, pos in mapeo_columnas.items():
                print(f"     • {nombre}: Col {pos + 1}")
            print(f"   Total columnas mapeadas: {len(mapeo_columnas)}")
            print("   " + "=" * 50)
        else:
            print(f"❌ DEBUG: No se pudieron detectar las columnas del template")

    # Mostrar estadísticas (incluyendo "Datos listos")
    total_facturas = df_consolidado["No. Factura"].nunique()
    total_conceptos = len(df_consolidado)

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("📄 Facturas", total_facturas)
    with col2:
        st.metric("📋 Conceptos", total_conceptos)
    with col3:
        st.metric("✅ Datos listos", "Sí" if template_listo else "No")

    # Mostrar el DataFrame
    st.dataframe(df_consolidado, use_container_width=True)

    # Botones de descarga CSV/Excel originales
    csv_buffer = df_consolidado.to_csv(index=False).encode("utf-8")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.download_button(
            label="📥 Descargar Datos (CSV)",
            data=csv_buffer,
            file_name="datos_facturas.csv",
            mime="text/csv",
        )

    with col2:
        excel_buffer = BytesIO()
        df_consolidado.to_excel(excel_buffer, index=False, engine="openpyxl")
        excel_buffer.seek(0)
        st.download_button(
            label="📥 Descargar Datos (Excel)",
            data=excel_buffer.getvalue(),
            file_name="datos_facturas.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # Botón Template SAT (solo si está listo)
    with col3:
        if template_listo:
            if st.button("📊 Generar Archivo SAT", type="primary"):
                try:
                    # Llenar el template con datos
                    wb_lleno = llenar_template_sat_con_datos(
                        wb, df_template, todas_facturas, fila_titulos, mapeo_columnas
                    )

                    if wb_lleno is None:
                        return

                    # Guardar en buffer para descarga
                    buffer = BytesIO()
                    wb_lleno.save(buffer)
                    buffer.seek(0)

                    st.success("🎉 ¡Archivo SAT generado correctamente!")
                    st.info("📝 Ya puedes descargar tu archivo para enviarlo al SAT")

                    # Botón de descarga
                    st.download_button(
                        label="📥 Descargar Archivo SAT",
                        data=buffer.getvalue(),
                        file_name="Template_SAT_Completo.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                except Exception as e:
                    st.error(
                        "❌ No se pudo generar el archivo. Contacta al equipo técnico."
                    )
                    print(f"❌ DEBUG ERROR TEMPLATE SAT: {str(e)}")
        else:
            st.warning("⚠️ Template SAT no disponible")


def mostrar_facturas(facturas, resumenes_hojas):
    """Muestra las facturas extraídas organizadas por hoja"""
    if not facturas:
        st.warning(
            "📋 No se encontraron facturas en este archivo. Verifica el formato."
        )
        return

    total_facturas = len(facturas)
    total_hojas = len(
        [
            h
            for h in resumenes_hojas.keys()
            if resumenes_hojas[h]["cantidad_facturas"] > 0
        ]
    )

    st.success(
        f"🎉 ¡Perfecto! Se encontraron {total_facturas} factura(s) en {total_hojas} hoja(s)"
    )

    # Mostrar resumen primero - AHORA CON DATOS REALES
    mostrar_resumen_hojas(resumenes_hojas, facturas)

    # **NUEVA SECCIÓN: Excel Consolidado**
    st.markdown("---")
    mostrar_excel_consolidado(facturas, resumenes_hojas)
    st.markdown("---")

    # Agrupar facturas por hoja
    facturas_por_hoja = {}
    for factura in facturas:
        nombre_hoja = factura["nombre_hoja"]
        if nombre_hoja not in facturas_por_hoja:
            facturas_por_hoja[nombre_hoja] = []
        facturas_por_hoja[nombre_hoja].append(factura)

    # Mostrar facturas organizadas por hoja
    st.subheader("📄 Facturas Detalladas por Hoja")

    # Volver a usar expanders por hoja, pero sin expanders por factura
    for nombre_hoja, facturas_hoja in facturas_por_hoja.items():
        with st.expander(
            f"📊 Hoja: {nombre_hoja} ({len(facturas_hoja)} facturas)", expanded=False
        ):
            # Mostrar información del cliente para esta hoja - USANDO DATOS REALES
            if facturas_hoja and facturas_hoja[0]["conceptos"]:
                primer_concepto = facturas_hoja[0]["conceptos"][0]
                st.subheader("Información del Cliente (de los datos procesados)")
                col1, col2, col3 = st.columns(3)

                with col1:
                    rfc_real = primer_concepto.get("RFC", "No encontrado")
                    st.write(f"**RFC:** {rfc_real}")

                with col2:
                    cliente_real = primer_concepto.get("CLIENTE", "No encontrado")
                    st.write(f"**Cliente:** {cliente_real}")

                with col3:
                    # Mostrar info original del header si existe
                    if facturas_hoja[0]["info_cliente"].get("CUENTA CONTABLE"):
                        st.write(
                            f"**Cuenta Contable:** {facturas_hoja[0]['info_cliente']['CUENTA CONTABLE']}"
                        )
                    else:
                        st.write("**Cuenta Contable:** No encontrada")

            st.markdown("---")

            # Mostrar cada factura directamente (SIN expanders por factura)
            for i, factura in enumerate(facturas_hoja, 1):
                st.subheader(f"🧾 Factura #{i} (Fila {factura['fila_rfc']})")

                # Mostrar conceptos
                st.write(f"**Conceptos:** {factura['total_conceptos']} elementos")

                if factura["conceptos"]:
                    # Convertir conceptos a DataFrame para mejor visualización
                    df_conceptos = pd.DataFrame(factura["conceptos"])
                    st.dataframe(df_conceptos, use_container_width=True)
                else:
                    st.info("📋 Esta factura no tiene conceptos registrados")

                # Separador entre facturas (excepto la última)
                if i < len(facturas_hoja):
                    st.markdown("---")


def cargar_archivo_excel(ruta_archivo):
    """Carga un archivo Excel y devuelve el objeto ExcelFile"""
    try:
        archivo_excel = pd.ExcelFile(ruta_archivo)
        return archivo_excel
    except Exception as e:
        st.error(
            f"❌ No se pudo abrir el archivo {ruta_archivo}. Verifica que no esté dañado."
        )
        return None


def mostrar_datos_excel(archivo_excel, nombre_archivo):
    """Muestra datos Excel parseando facturas de todas las hojas"""
    if archivo_excel is None:
        return

    nombres_hojas = archivo_excel.sheet_names

    # Parsear todas las hojas a la vez
    st.subheader(f"📊 Facturas Parseadas de todas las hojas en: {nombre_archivo}")
    todas_facturas, resumenes_hojas = extraer_todas_facturas(archivo_excel)
    mostrar_facturas(todas_facturas, resumenes_hojas)

    # Mostrar resumen total
    total_facturas = len(todas_facturas)
    total_hojas = len(nombres_hojas)
    st.info(
        f"**Resumen Total:** {total_facturas} facturas procesadas de {total_hojas} hoja(s)"
    )


def main():
    st.set_page_config(
        page_title="Analizador de Facturas Hanova",
        page_icon="🧾",
        layout="wide",
        initial_sidebar_state="collapsed",
    )

    st.title("🧾 Analizador de Facturas Hanova")
    st.markdown("Sube archivos Excel para parsear facturas de **todas las hojas**")
    st.markdown(
        "💡 *La aplicación procesa todas las hojas en cada archivo Excel y maneja diferentes clientes por hoja*"
    )
    st.markdown("---")

    # Sección de subida de archivos
    st.header("📤 Subir Archivos Excel")
    archivos_subidos = st.file_uploader(
        "Elige archivos Excel",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        help="Sube archivos Excel que contengan facturas con CLIENTE en A3 y marcadores RFC. Todas las hojas serán procesadas automáticamente.",
    )

    # Procesando archivos subidos
    if archivos_subidos:
        st.subheader("Análisis de Archivos Subidos")

        pestañas = st.tabs([f"📄 {archivo.name}" for archivo in archivos_subidos])

        for pestaña, archivo_subido in zip(pestañas, archivos_subidos):
            with pestaña:
                try:
                    # Guardar archivo subido en ubicación temporal
                    with tempfile.NamedTemporaryFile(
                        delete=False, suffix=".xlsx"
                    ) as archivo_temp:
                        archivo_temp.write(archivo_subido.getvalue())
                        ruta_archivo_temp = archivo_temp.name

                    # Cargar y mostrar el archivo
                    datos_excel = cargar_archivo_excel(ruta_archivo_temp)

                    # Mostrar datos parseados (siempre parsear facturas)
                    mostrar_datos_excel(datos_excel, archivo_subido.name)

                    # Limpiar archivo temporal
                    os.unlink(ruta_archivo_temp)

                except Exception as e:
                    st.error(
                        f"❌ No se pudo procesar {archivo_subido.name}. Verifica que sea un archivo Excel válido."
                    )
    else:
        st.info("👆 Sube uno o más archivos Excel con facturas para comenzar")


if __name__ == "__main__":
    main()
